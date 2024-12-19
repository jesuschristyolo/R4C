import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import timedelta
from django.utils import timezone
from django.http import HttpResponse
from robots.models import Robot


def generate_week_report():
    today = timezone.now()
    last_week = today - timedelta(days=7)

    robot_models = Robot.objects.filter(created__gte=last_week).values_list('model', flat=True).distinct()

    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if not robot_models:
        wb.create_sheet(title="Роботов найдено не было")

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')

    for model in robot_models:

        robots = Robot.objects.filter(model=model, created__gte=last_week)

        sheet = wb.create_sheet(title=f"model {model}")
        sheet.append(['Модель', 'Версия', 'Количество за неделю'])

        for col in range(1, 4):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment

        versions = robots.values('version').distinct()
        row = 2
        for version_entry in versions:
            version = version_entry['version']
            count = robots.filter(version=version).count()
            sheet.append([model, version, count])
            row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=weekly_robot_report.xlsx'

    wb.save(response)
    return response
